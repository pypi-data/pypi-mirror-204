# -*- coding: utf-8 -*-
"""
This module offers a set of operations that user can modify their excel files.
"""
import pathlib
import warnings

try:
    from importlib.resources import files
except ImportError:
    from importlib_resources import files

from os import environ, remove
from pathlib import Path
from typing import List, Optional, Tuple, Union

import openpyxl
from dotenv import load_dotenv
from openpyxl.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet
from urllib3 import disable_warnings

from .excel_definition import ExcelDefinition
from .jira_client import JiraClient
from .sprint_schedule import SprintScheduleStore
from .story import (
    Story,
    StoryFactory,
    sort_stories_by_inline_weights,
    sort_stories_by_property_and_order,
    sort_stories_by_raise_ranking,
)

__all__ = [
    "read_excel_file",
    "output_to_excel_file",
    "output_to_csv_file",
    "run_steps_and_sort_excel_file",
]

# Currently, the openpyxl package will report an obsolete warning.
warnings.simplefilter(action="ignore", category=UserWarning)
# Disable the HTTPS certificate verification warning.
disable_warnings()

HERE = pathlib.Path(__file__).resolve().parent
ASSETS = HERE / "assets"


def read_excel_file(
    file: Union[str, Path],
    excel_definition: ExcelDefinition,
    sprint_schedule: SprintScheduleStore,
) -> Tuple[List[str], List[Story]]:
    """
    Read and parse the excel file

    :parm file:
        The excel file that you want to read

    :parm excel_definition:
        The excel column definition which is imported
        from the :py:class:`ExcelDefinition`

    :parm sprint_schedule:
        The priority mapping for the :py:class:`Milestone` object.

    :return:
        A :py:class:`tuple` object which contains a list of column
        name and a list of :py:class:`Story`.
    """

    if file is None or not pathlib.Path(file).is_absolute():
        raise ValueError("The input excel file is invalid.")

    if not pathlib.Path(file).exists():
        raise ValueError(f"The input excel file: {file} cannot be found.")

    with open(str(file), mode="rb") as raw_file:
        work_book: Workbook = openpyxl.load_workbook(
            raw_file, read_only=True, keep_vba=False, data_only=True, keep_links=True
        )

        if work_book.active is None or (
            not isinstance(work_book.active, Worksheet)
            and not isinstance(work_book.active, ReadOnlyWorksheet)
        ):
            work_book.close()
            raise ValueError("The input excel file doesn't contain any sheets.")

        sheet: Union[Worksheet, ReadOnlyWorksheet] = work_book.active

        # The count of column is taking from the definition file to avoid too
        # many columns inside the excel file. Also, need to avoid exceed
        # the range of the actual count.
        column_count = min(excel_definition.max_column_index, sheet.max_column)

        if sheet.max_row < 2:
            work_book.close()
            return ([], [])

        columns: List[str] = []

        for column_index in range(1, column_count + 1):
            columns.append(str(sheet.cell(row=1, column=column_index).value))

        stories = []

        excel_definition_columns = excel_definition.get_columns()
        story_factory = StoryFactory(excel_definition_columns)

        for row in sheet.iter_rows(
            min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(columns)
        ):
            if _should_skip(row):
                continue

            story: Story = story_factory.create_story()
            for column_index, column in enumerate(row):
                definition_column = excel_definition_columns[column_index]
                if definition_column["name"] is None:
                    continue
                story.set_value(
                    definition_column["type"], definition_column["name"], column.value
                )

            story.calc_sprint_schedule(sprint_schedule)
            stories.append(story)

        work_book.close()
        raw_file.close()
    return (columns, stories)


def _should_skip(row: tuple) -> bool:
    if len(row) == 0:
        return True
    first_cell_value = row[0].value
    if first_cell_value is None or len(str(first_cell_value)) == 0:
        return True
    return False


def output_to_csv_file(
    file: Union[str, Path],
    stories: "List[Story]",
    over_write: bool = True,
):
    if file is None or not pathlib.Path(file).is_absolute():
        raise ValueError("The file is invalid.")

    if pathlib.Path(file).exists():
        if over_write is True:
            remove(file)
        else:
            raise ValueError(f"The csv file: {file} is already exist.")

    with open(file, mode="x+", encoding="utf-8") as csv_file:
        for story in stories:
            csv_file.writelines([str(story)])


def output_to_excel_file(
    file: Union[str, Path],
    stories: "List[Story]",
    excel_definition: ExcelDefinition,
    columns_in_excel: Optional[List[str]] = None,
    over_write: bool = True,
):
    """
    Generate excel file

    :parm file:
        Output excel file name including the path

    :parm stories:
        A list of :py:class:`Story` which need to be wrote to the excel

    :parm excel_definition:
        The excel column definition which is imported from
        the :py:class:`ExcelDefinition`

    :parm columns_in_excel:
        Using separate column names instead of importing from
        the :py:class:`ExcelDefinition`. Usually, it comes from the
        input excel file.

    :parm over_write:
        Whether or not the exist output file will be over-write.
    """
    if file is None or not pathlib.Path(file).is_absolute():
        raise ValueError("The output file name is invalid.")

    if pathlib.Path(file).exists():
        if over_write is True:
            try:
                remove(file)
            except PermissionError as e:
                print(e)
                return
        else:
            raise ValueError(f"The output excel file: {file} is already exist.")

    work_book = openpyxl.Workbook(write_only=False)

    if work_book.active is None or (
        not isinstance(work_book.active, Worksheet)
        and not isinstance(work_book.active, Worksheet)
    ):
        work_book.close()
        raise ValueError("The output excel file cannot be generated.")

    sheet: Worksheet = work_book.active

    excel_definition_columns = excel_definition.get_columns()

    # Use original excel column name first.
    columns = columns_in_excel
    if columns is None:
        columns = [column["name"] for column in excel_definition_columns]

    for column_index, column in enumerate(columns):
        cell = sheet.cell(row=1, column=column_index + 1)
        # There are three kinds of Cells. Only the Cell has the value attribute.
        if hasattr(cell, "value"):
            setattr(cell, "value", column)

    if stories is not None and stories:
        for row_index, story in enumerate(stories):
            for column in excel_definition_columns:
                if column["name"] is None:
                    continue
                cell = sheet.cell(row=row_index + 2, column=column["index"])
                if hasattr(cell, "value"):
                    setattr(cell, "value", story.format_value(column["name"]))

    work_book.save(str(file))
    work_book.close()


def _query_jira_information(
    stories: List[Story], excel_definition: ExcelDefinition
) -> bool:
    load_dotenv(ASSETS / ".env")

    jira_url: Optional[str] = environ.get("JIRA_URL", default=None)
    if jira_url is None or jira_url.isspace() or len(jira_url) == 0:
        print(
            "The jira url is invalid. Please use the update-jira-info command to add/update url."
        )
        return False

    jira_acccess_token: Optional[str] = environ.get("JIRA_ACCESS_TOKEN", default=None)
    if (
        jira_acccess_token is None
        or jira_acccess_token.isspace()
        or len(jira_acccess_token) == 0
    ):
        print(
            "The jira access token is invalid. Please use the update-jira-info command to add/update token."
        )
        return False

    jira_client = JiraClient(jira_url, jira_acccess_token)

    if not jira_client.health_check():
        print(
            "The jira access token is revoked. Please use the update-jira-info command to add/update token."
        )
        return False

    print(f"Jira info: {jira_url}")

    jira_fields = []

    for definition_column in excel_definition.get_columns():
        if (
            definition_column["name"] is None
            or definition_column["jira_field_mapping"] is None
        ):
            continue
        jira_fields.append(
            {
                "name": definition_column["name"],
                "jira_name": definition_column["jira_field_mapping"]["name"],
                "jira_path": definition_column["jira_field_mapping"]["path"],
            }
        )

    jira_query_result = jira_client.get_stories_detail(
        [story["storyId"].strip() for story in stories], jira_fields
    )

    for story in stories:
        story_id: str = story["storyId"].lower().strip()
        if story_id in jira_query_result:
            for jira_field in jira_fields:
                story[jira_field["name"]] = jira_query_result[story_id][
                    jira_field["jira_name"]
                ]
        else:
            # Story ID has been changed because of convertion.
            temp_result = jira_client.get_stories_detail([story_id], jira_fields)
            if len(temp_result) > 0:
                story["storyId"] = list(temp_result.keys())[0].upper()
                for jira_field in jira_fields:
                    story[jira_field["name"]] = list(temp_result.values())[0][
                        jira_field["jira_name"]
                    ].upper()
                print(
                    f"Story id has been changed. Previous: {story_id.upper()}, Current: {story['storyId'].upper()}"
                )
            else:
                print(f"Cannot find related information for story: {story_id}")
                story.need_sort = False
                continue

    return True


def run_steps_and_sort_excel_file(
    input_file: Union[str, Path],
    output_file: Union[str, Path],
    excel_definition_file: Optional[Union[str, Path]] = None,
    sprint_schedule_file: Optional[Union[str, Path]] = None,
    over_write: bool = True,
):
    """
    Sort the excel file and output the result

    :parm input_file:
        The excel file need to be sorted. (Absolute path only)

    :parm output_file:
        The sorted excel file location. (Absolute path only)

    :parm sprint_schedule_file:
        The JSON file which contains the priority list to calculate the :py:class:`Milestone`

    :parm excel_definition_file:
        The JSON file which contains the input excel file's structure.

    :parm over_write:
        Whether or not the exist output file will be over-write.
    """
    sprint_schedule = SprintScheduleStore()
    if sprint_schedule_file is None:
        sprint_schedule.load(
            files("jira_assistant.assets")
            .joinpath("sprint_schedule.json")
            .read_text(encoding="utf-8")
        )
        print("Using default sprint schedule...")
    else:
        sprint_schedule.load_file(sprint_schedule_file)
        print("Using custom sprint schedule...")

    excel_definition = ExcelDefinition()
    if excel_definition_file is None:
        excel_definition.load(
            files("jira_assistant.assets")
            .joinpath("excel_definition.json")
            .read_text(encoding="utf-8")
        )
        print("Using default excel definition...")
    else:
        excel_definition.load_file(excel_definition_file)
        print("Using custom excel definition...")

    validation_result = excel_definition.validate()
    if len(validation_result) != 0:
        print(
            "Validating excel definition failed. Please check below information to fix first."
        )
        for item in validation_result:
            print(item)
        return
    print("Validating excel definition success.")

    excel_columns, stories = read_excel_file(
        input_file, excel_definition, sprint_schedule
    )

    if stories is None or len(stories) == 0:
        print("There are no stories inside the excel file.")
        return

    # Execute pre-process steps
    pre_process_steps = excel_definition.get_pre_process_steps()

    for pre_process_step in pre_process_steps:
        print(f"Executing step: {pre_process_step['name']}...")
        if pre_process_step["name"].lower() in "RetrieveJiraInformation".lower():
            need_call_jira_api: bool = False
            for excel_definition_column in excel_definition.get_columns():
                if excel_definition_column["jira_field_mapping"] is not None:
                    need_call_jira_api = True
                    break

            if need_call_jira_api:
                stories_need_call_jira: List[Story] = []
                for story in stories:
                    if story.need_sort:
                        stories_need_call_jira.append(story)
                if not _query_jira_information(
                    stories_need_call_jira, excel_definition
                ):
                    print("Retrieve jira information failed.")
                    return
        elif pre_process_step["name"].lower() in "FilterOutStoryWithoutId".lower():
            for story in stories:
                if story["storyId"] is None:
                    story.need_sort = False
        elif (
            pre_process_step["name"].lower()
            in "FilterOutStoryBasedOnJiraStatus".lower()
        ):
            for story in stories:
                if story["status"] is not None and story[
                    "status"
                ].upper() in pre_process_step["config"].get("JiraStatuses", []):
                    story.need_sort = False
        print("Executing finish.")

    stories_no_need_sort = []
    stories_need_sort = []

    for story in stories:
        if story.need_sort:
            stories_need_sort.append(story)
        else:
            stories_no_need_sort.append(story)

    # Execute sorting logic.
    sort_strategies = excel_definition.get_sort_strategies()

    for sort_strategy in sort_strategies:
        print(f"Executing {sort_strategy['name']} sorting...")
        if sort_strategy["name"] is None:
            continue
        if sort_strategy["name"].lower() in "InlineWeights".lower():
            stories_need_sort = sort_stories_by_inline_weights(stories_need_sort)
        elif sort_strategy["name"].lower() in "SortOrder".lower():
            sort_stories_by_property_and_order(
                stories_need_sort, excel_definition, sort_strategy["config"]
            )
        elif sort_strategy["name"].lower() in "RaiseRanking".lower():
            stories_need_sort = sort_stories_by_raise_ranking(
                stories_need_sort, excel_definition, sort_strategy["config"]
            )
        print("Executing finish.")

    output_to_excel_file(
        output_file,
        stories_need_sort + stories_no_need_sort,  # First output the sorted stories.
        excel_definition,
        excel_columns,
        over_write,
    )

    print(f"{output_file} has been saved.")
