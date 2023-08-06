"""
This file is used to read out data from specific dataset.
The dataset path is defined in test_configuration/config.yaml.

DatasetHandler is the super class for all project dataset,
the implement can reference to the  ford_dataset_handler.FordDatasetHandler

External parameter:
1. excel_data: All excel data from the specify path. The data structure: {sheet_name: sheet_value}

External function:
1.get_row_column_array/2: Get the part of data(row_start, row_end, column_start, column_end)
                          from the specify sheet name.
2.merge_data_according_row/2: merge the dict according to the same row
3.get_row_array/5: get the data format with row
4.get_column_array/5: get the data format with column

@Author: Siwei.Lu
@Date: 2022.11.26
"""
import sys
import enum

import openpyxl

from .. import logger


class DatasetHandler:
    def __init__(self, path: str, used_sheet: enum):
        self.__excel_data = self._get_excel_data(path, used_sheet)

    @staticmethod
    def _get_excel_data(path, used_sheet: enum):
        """
        Get excel data from specific path.
        :param path: str
        :return: Dict: {sheet_name: sheet_value}
        """
        sheet_data_dict = {}

        try:
            data = openpyxl.load_workbook(path)
            data.close()
        except BaseException as err:
            print("Dataset open failed! \n{}".format(err))
            logger.critical("Dataset open failed! \n{}".format(err))
            sys.exit()

        for key, value in used_sheet.__members__.items():
            sheet = value.value
            if sheet not in data.sheetnames:
                logger.critical("The sheet named {} is not exist in the file {}".format(sheet, path))
                raise Exception("The sheet named {} is not exist in the file {}".format(sheet, path))
            sheet_data_dict[sheet] = data[sheet]

        logger.info("Get Dataset File Successful from path: {}".format(path))
        print("Get Dataset File Successful !")

        return sheet_data_dict

    def _get_row_array(self, sheet_name, row_start, row_end, column_start, column_end):
        """
        Return the data according to given row and column.
        Row and column are range.
        :param sheet_name: str
        :param row_start: int
        :param row_end: int
        :param column_start: int
        :param column_end: int
        :return: data_sort_according_row: Dict: {Row, Row_value}
        """
        data_sort_according_row = {}
        for i in range(row_start, row_end + 1):
            column_datas = []
            for j in range(column_start, column_end + 1):
                column_datas.append(self.__excel_data[sheet_name].cell(row=i, column=j).value)

            data_sort_according_row[i] = column_datas
        return data_sort_according_row

    def _get_column_array(self, sheet_name, row_start, row_end, column_start, column_end):
        """
        Return the data according to given row and column.
        Row and column are range.
        :param sheet_name: str
        :param row_start: int
        :param row_end: int
        :param column_start: int
        :param column_end: int
        :return: data_sort_according_col: Dict: {Column, Column_value}
        """
        data_sort_according_col = {}
        for j in range(column_start, column_end + 1):
            row_datas = []
            for i in range(row_start, row_end + 1):
                row_datas.append(self.__excel_data[sheet_name].cell(row=i, column=j).value)

            data_sort_according_col[j] = row_datas

        return data_sort_according_col

    @staticmethod
    def merge_data_according_row(dict1, dict2):
        for row, row_value in dict1.items():
            if dict2[row]:
                dict1[row] = row_value + dict2[row]
        return dict1

    def get_row_column_array(self, sheet_name, list_info: list, based_on_row_or_col=0):
        """
        @param based_on_row_or_col: 0: based on column, 1: based on row
        @param list_info: [row_start, row_end, column_start, column_end]
        @param sheet_name:
        @return: list
        """
        row_start = int(list_info[0])
        row_end = int(list_info[1])
        column_start = int(list_info[2])
        column_end = int(list_info[3])

        array_1 = []
        if (row_start == row_end) and (column_start == column_end):
            return_value = self.__excel_data[sheet_name].cell(row=row_start, column=column_end).value
        elif row_end == row_start:
            for j in range(column_start, column_end + 1):
                array_1.append(self.__excel_data[sheet_name].cell(row=row_start, column=j).value)
            return_value = array_1
        elif column_start == column_end:
            for j in range(row_start, row_end + 1):
                array_1.append(self.__excel_data[sheet_name].cell(row=j, column=column_end).value)
            return_value = array_1
        else:
            if based_on_row_or_col == 0:
                for j in range(column_start, column_end + 1):
                    array_2 = []
                    for i in range(row_start, row_end + 1):
                        array_2.append(self.__excel_data[sheet_name].cell(row=i, column=j).value)
                    array_1.append(array_2)
                return_value = array_1
            else:
                for j in range(row_start, row_end + 1):
                    array_2 = []
                    for i in range(column_start, column_end + 1):
                        array_2.append(self.__excel_data[sheet_name].cell(row=j, column=i).value)
                    array_1.append(array_2)
                return_value = array_1

        return return_value

    @property
    def excel_data(self):
        return self.__excel_data
