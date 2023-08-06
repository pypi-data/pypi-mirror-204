
# ---------------------------------------------------------------------------------------------------------------
# PyxlSQL project
# This program and library is licenced under the European Union Public Licence v1.2 (see LICENCE)
# developed by fabien.battini@gmail.com
# ---------------------------------------------------------------------------------------------------------------

import re
import os
import importlib
import openpyxl
import openpyxl.styles
from PyxlSql.pyxlErrors import PyxlSqlSheetError, PyxlSqlError, PyxlSqlExecutionError, PyxlSqlInternalError
from PyxlSql.pyxlSheets import PyxlSheet

# ---------------------------------------------------------------------------------------------------
# class NamedWB
# ---------------------------------------------------------------------------------------------------


class PyxlWorkbook:
    """
    class to manage excel workbooks structured as a database
    """
    def __init__(self, file_name, create=False, first_row=1, first_column= 1, last_row=None, last_column=None,
                 font=None, file_path: list=[]):
        self.file_path = file_path or []
        self.filename = self.find_file(file_name)

        if self.filename is None:
            print(f"FATAL ERROR: Cannot find '{file_name}' in path '{self.file_path}', aborting")
            exit(-1)

        local_path = os.path.dirname(self.filename)
        if local_path not in self.file_path:
            self.file_path.append(local_path)

        try:
            self.wb: openpyxl.workbook = openpyxl.load_workbook(filename=self.filename)
        except OSError as error:
            print(f"FATAL ERROR: Cannot open '{file_name}' : {str(error)}, aborting")
            exit(-1)

        self.sheets = {}  # a dictionary  string --> NamedWS

        self.book_default_font = openpyxl.styles.Font(name='Century Gothic', size=11) if font is None else font
        self.wbs = {}       # all workbooks referenced by this one
        self.imported = {}  # dictionary of imported symbols, will be used when eval is called
        self.sheets_to_delete = []
        self.import_module("functools")
        self.theme = self.init_theme()
        self.first_row = first_row
        self.last_row = last_row
        self.first_column = first_column
        self.last_column = last_column
        if create:
            self.wb = openpyxl.Workbook()
            return


    # <a:dk1><a:sysClr val="windowText" lastClr="000000"/></a:dk1>
    # <a:lt1><a:sysClr val="window" lastClr="FFFFFF"/></a:lt1>
    # <a:dk2><a:srgbClr val="44546A"/></a:dk2>                   # noqa
    # <a:lt2><a:srgbClr val="E7E6E6"/></a:lt2>                   # noqa
    # <a:accent1><a:srgbClr val="5B9BD5"/></a:accent1>           # noqa
    # <a:accent2><a:srgbClr val="ED7D31"/></a:accent2>           # noqa
    # <a:accent3><a:srgbClr val="A5A5A5"/></a:accent3>           # noqa
    # <a:accent4><a:srgbClr val="FFC000"/></a:accent4>           # noqa
    # <a:accent5><a:srgbClr val="4472C4"/></a:accent5>           # noqa
    # <a:accent6><a:srgbClr val="70AD47"/></a:accent6>           # noqa
    # <a:hlink><a:srgbClr val="0563C1"/></a:hlink>               # noqa
    # <a:folHlink><a:srgbClr val="954F72"/></a:folHlink>         # noqa
    #
    def init_theme(self):
        theme = []
        theme_desc = self.wb.loaded_theme.decode('utf-8')

        m = re.search(r'<a:lt1><a:sysClr val=".*" lastClr="([\dABCDEF]*)"/></a:lt1>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")
        # CAVEAT: for some strange reason, the first 4 items do not appear in the appropriate order inside the xml
        # we have to reorder to get the Windows file color scheme !

        m = re.search(r'<a:dk1><a:sysClr val=".*" lastClr="([\dABCDEF]*)"/></a:dk1>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:lt2><a:srgbClr val="([\dABCDEF]*)"/></a:lt2>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:dk2><a:srgbClr val="([\dABCDEF]*)"/></a:dk2>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")


        m = re.search(r'<a:accent1><a:srgbClr val="([\dABCDEF]*)"/></a:accent1>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:accent2><a:srgbClr val="([\dABCDEF]*)"/></a:accent2>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:accent3><a:srgbClr val="([\dABCDEF]*)"/></a:accent3>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:accent4><a:srgbClr val="([\dABCDEF]*)"/></a:accent4>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:accent5><a:srgbClr val="([\dABCDEF]*)"/></a:accent5>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:accent6><a:srgbClr val="([\dABCDEF]*)"/></a:accent6>', theme_desc)
        if m:
            theme.append(m.group(1))
        return theme

    def find_file(self, filename):
        if filename is None:
            return None
        if os.path.exists(filename):
            return filename
        for d in self.file_path:
            if d[-1] != "/":
                d += "/"
            f = os.path.realpath(d + filename)
            if os.path.exists(f):
                return f
        return None

    def get_sheet(self, sheet_name, first_row=None, first_column=None, last_row=None, last_column=None,
                  raise_exception=True, to_read=True):
        first_row = first_row or self.first_row

        if sheet_name is None:
            return None
        if sheet_name in self.sheets:
            return self.sheets[sheet_name]

        if sheet_name in self.wb:
            sheet = self.wb[sheet_name]
            self.sheets[sheet.title] = PyxlSheet(self, sheet, to_read=to_read,
                                                 first_row=first_row,last_row=last_row,
                                                 first_column=first_column,last_column=last_column)
            return self.sheets[sheet_name]

        m = re.search("\"?\'?([^\"\'[]+)\"?\'?(\\[(.*)])?", sheet_name)
        if m is not None:
            first_name: str = m.group(1)
            second_name: str = m.group(3)
            if first_name is not None and second_name is not None:
                workbook = self.get_workbook(first_name)
                if workbook is not None:
                    if second_name in workbook.sheets:
                        return workbook.sheets[second_name]
                    if second_name in workbook.wb.sheetnames:
                        sheet = workbook.wb[second_name]

                        workbook.sheets[sheet.title] = PyxlSheet(self, sheet, to_read=to_read,
                                                                 first_row=first_row,last_row=last_row,
                                                                 first_column=first_column,last_column=last_column)
                        return workbook.sheets[second_name]

            # here, we have not found the sheet

        if raise_exception:
            raise PyxlSqlSheetError(sheet_name, "workbook")
        return None

    def delete_sheet(self, sheet_name):
        """Deletes the sheet from the workbook"""
        sheet = self.get_sheet(sheet_name)
        if sheet is not None:
            del self.wb[sheet_name]

    def rename_sheet(self, old_sheet_name:str, new_sheet_name:str):
        """Rename the sheet from the workbook"""
        sheet = self.get_sheet(old_sheet_name)
        # sheet cannot be None, get_sheet would have raised an exception

        sheet.rename(new_sheet_name)
        self.sheets[new_sheet_name]=sheet
        self.sheets.pop(old_sheet_name)

    def save(self, file_name):
        current_dir = os.path.dirname(os.path.realpath(self.filename))
        try:
            self.wb.save(current_dir + "\\" + file_name)
        except OSError as err:
            raise PyxlSqlExecutionError(f" file '{current_dir}\\{file_name}' is read-only", str(err))

    @staticmethod
    def local_open(file_name: str, mode: str):
        try:
            ins = open(file_name, mode)
        except OSError as error:
            raise PyxlSqlExecutionError(f"FATAL ERROR: Cannot open('{file_name}',{mode})", str(error))
        return ins

    @staticmethod
    def escape(excel_str: str):
        html_str = excel_str.encode('ascii', 'xmlcharrefreplace')
        html_str =  html_str.decode('utf-8')
        return html_str

    def export_html(self, file_name: str, css: str=""):

        outs = self.local_open(file_name, "w")

        outs.writelines('<!DOCTYPE html>\n')
        outs.writelines('<html>\n')
        outs.writelines('<head>\n')
        outs.writelines('    <meta http-equiv=Content-Type content="text/html; charset=windows-1252">\n')
        outs.writelines(f'    <title>{self.filename}</title>\n')
        if css == "":
            outs.writelines('   <style>\n')
            outs.writelines('     html, body, h1, h2, h3, p { margin: 5; padding: 5; border: 0; \n')
            outs.writelines('            font-size: 100%; font-family: inherit; vertical-align: baseline; }\n')
            outs.writelines('     table, td, th { margin: 0; padding: 1; border: 1px solid black;\n')
            outs.writelines('            border-collapse: collapse; vertical-align: top;}\n')
            outs.writelines('     th  {text-align: left;}\n')

            outs.writelines('     H1 {background: #f8f8f8; width: 100%; border-bottom: 1px solid #ccc;  \n')
            outs.writelines('            border-top: 1px solid #ccc; font-size: 1.25em; display: inline-block;}\n')
            outs.writelines('     H2 {background: #888888; width: 100%; margin-top: .5em; font-size: 1em; display: inline-block;}\n')
            outs.writelines('   </style>\n')
        else:
            outs.writelines(f'    <link rel="stylesheet" href="{css}" type="text/css">\n')


        outs.writelines('</head>\n')
        outs.writelines('<body>\n')
        outs.writelines(f'    <h1>{self.filename}</h1>\n')

        for cur_sheet in self.sheets.values():
            outs.writelines('   <div  class="sheet">\n')
            outs.writelines(f'      <h2>Sheet [{cur_sheet.get_name()}]</h2>\n')
            outs.writelines('      <br><br>\n')
            outs.writelines('      <table>\n')
            outs.writelines('        <thead>\n')                                           # noqa
            outs.writelines('          <tr><th>1</th>\n')
            for col_name in cur_sheet.columns:
                outs.write(f'            <th {cur_sheet.get_html_style(1, col_name)}>{col_name}</th>\n')
            outs.write('          </tr>\n')
            outs.write('        </thead>\n')                                               # noqa

            for lineno in cur_sheet.get_row_range():
                outs.write(f'         <tr><td>{lineno}</td>\n')
                for col_name in cur_sheet.columns:
                    outs.write(f'           <td {cur_sheet.get_html_style(lineno, col_name)}>')
                    outs.write(f'               {self.escape(cur_sheet.get_html_value(lineno, col_name))}</td>\n')
                outs.write('         </tr>\n')
            outs.write('      </table>\n')
            outs.write('   </div>\n')
        outs.write('</body>\n')
        outs.write('</html>\n')
    def get_workbook(self, name):
        if name not in self.wbs:
            self.wbs[name] = PyxlWorkbook(name, file_path=self.file_path)
        return self.wbs[name]

    def import_module(self, module, sub_modules=None):
        mod = importlib.import_module(module)
        if sub_modules is None:
            self.imported[module] = mod
            return

        if sub_modules == '*':
            if hasattr(mod, '__all__'):
                item_list = mod.__all__
            else:
                raise PyxlSqlError(f"ERROR: IMPORT {module} SUBS {sub_modules} : ABORTING the import",
                                   "        The module does not contain a __all__ symbol that allows importing *")
        else:
            item_list = sub_modules.split()

        for item in item_list:
            self.imported[item] = getattr(mod, item)
