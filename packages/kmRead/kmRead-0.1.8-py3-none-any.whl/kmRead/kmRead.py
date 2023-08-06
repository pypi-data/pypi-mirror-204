# encoding=utf8
import csv
import json
import os
import sys
from openpyxl import Workbook,styles
import zipfile


def un_zxm(file_name):
    """unzip zip file"""
    zip_file = zipfile.ZipFile(file_name)
    data=zip_file.read('content.json')
    data=json.loads(data)
    zip_file.close()
    return data

def write_csv_file(path, head, data):
    '''
    用途：把list数据写入csv文件中
    path 文件地址
    head 标题行内容，格式list
    data 正式内容，格式list
    '''
    try:
        with open(path, 'w', newline='', encoding='utf8') as csv_file:
            writer = csv.writer(csv_file, dialect='excel')
            if head is not None:
                writer.writerow(head)
            for row in data:
                writer.writerow(row)
            print("Write a CSV file to path %s Successful." % path)
    except Exception as e:
        print("Write an CSV file to path: %s, Case: %s" % (path, e))

def write_excel_file(path, head, data):
    '''
    用途：把list数据写入csv文件中
    path 文件地址
    head 标题行内容，格式list
    data 正式内容，格式list
    '''
    try:
        wb = Workbook()
        ws = wb.active
        ws.column_dimensions['B'].width=100
        ws.column_dimensions['E'].width=100
        ws.column_dimensions['F'].width=50
        for indexV in range(len(head)):
            ws.cell(1,indexV+1,head[indexV])
        for indexh in range(len(data)):
            for indexl in range(len(data[indexh])):
                ws.cell(indexh+2,indexl+1,data[indexh][indexl])
                if (indexl==5) :
                    ws['F'+str(indexh+2)].alignment = styles.Alignment(wrapText=True)
        wb.save(path)
        print("Write a excel file to path %s Successful." % path)
    except Exception as e:
        print("Write an excel file to path: %s, Case: %s" % (path, e))


def toFile(xuid, pid, data, fileName,isCSV=True):
    '''
    用途：把数据整理成list，写入csv文件中
    xuid 需求id 
    pid p账号
    '''
    datas = []
    csvList = ['用例目录', '用例名称', '需求ID', '前置条件', '用例步骤', '预期结果', '用例类型', '用例状态',
               '用例等级', '创建人', '是否实现自动化', '是否上架', '自动化测试类型', '自动化测试平台', '移动系统', '是否自动化', '标签']
    for k, v in data.items():
        tmpList = []
        tmpDict = {'用例名称': str(k)[1:], '需求ID': str(
            xuid), '用例步骤': k[1:], '预期结果': v, '用例类型': '功能测试', '用例状态': '正常', '用例等级': '中', '创建人': str(pid)}
        for vdata in csvList:
            if vdata in tmpDict:
                tmpList.append(tmpDict.get(vdata))
            else:
                tmpList.append("")
        datas.append(tmpList.copy())
    if(isCSV):
        write_csv_file(fileName.replace('.xmind', '.csv').replace('.km', '.csv').replace('.zxm','.csv'), csvList, datas)
    else:
        write_excel_file(fileName.replace('.xmind', '.xlsx').replace('.km', '.xlsx').replace('.zxm','.xlsx'), csvList, datas)

def getDataFromKM(root, data, nowList):
    '''
    用途：使用嵌套循环获取km文件内容并生产list数据
    root 文件目录
    data 获取单层内容
    nowList 当前已获取的数据
    '''
    tmpData = str(data['data']['text']).strip()  # 临时数据报保存当前值
    if len(data['children']) > 0:  # 判断是否含有子目录
        for childrenData in data['children']:
            getDataFromKM(root+','+tmpData, childrenData, nowList)
    else:
        nowList.append({root: tmpData})
    return nowList

def getDataFromXmind(root, data, nowList):
    '''
    用途：使用嵌套循环获取km文件内容并生产list数据
    root 文件目录
    data 获取单层内容
    nowList 当前已获取的数据
    '''
    tmpData = str(data['title']).strip()  # 临时数据报保存当前值
    if 'children' in data and len(data['children']['attached']) > 0:  # 判断是否含有子目录
        for childrenData in data['children']['attached']:
            getDataFromXmind(root+','+tmpData, childrenData, nowList)
    else:
        nowList.append({root: tmpData})
    return nowList

def getFile(path):
    '''
    用途：获取文件并转成dict格式
    path 文件路径
    '''
    if not os.path.exists(path):
        print("文件不存在："+str(path))
        sys.exit(1)
    data = {}  # 存数据用
    with open(path, encoding='utf8') as f:
        data = f.read()
    return json.loads(data, encoding='utf8')

def getDataFromZXM(root, data, nowList):
    '''
    用途：使用嵌套循环获取zxm文件内容并生产list数据
    root 文件目录
    data 获取单层内容
    nowList 当前已获取的数据
    '''

    tmpData = str(data['data']['text']).strip().strip('\n').strip(r'\n')  # 临时数据报保存当前值
    if len(data['children']) > 0:  # 判断是否含有子目录
        for childrenData in data['children']:
            getDataFromKM(root+','+tmpData, childrenData, nowList)
    else:
        nowList.append({root: tmpData})
    return nowList



def run(filepath, xuid, pid,isCSV=True):
    '''
    用途：执行脚本
    filepath km文件的绝对路径
    xuid 需求id 
    pid p账号
    '''
    if str(filepath).endswith('.km'):
        fileLine = getDataFromKM('', getFile(filepath)['root'], [])
    elif str(filepath).endswith('.zxm'):
        fileLine = getDataFromZXM('', un_zxm(filepath)['root'], [])
    elif str(filepath).endswith('.xmind'):
        fileLine = getDataFromXmind('', un_zxm(filepath)[0]['rootTopic'], [])
    else:
        raise Exception("文件格式错误")
    caseDict = {}
    for line in fileLine:
        for k, v in line.items():
            if caseDict.get(k) is None:
                caseDict[k]=v
            else:
                caseDict[k]+="\n"+v
    toFile(str(xuid), pid, caseDict, filepath,isCSV)

def put():
    f=""
    if (len(sys.argv)==1):
        f=input("请填写文件地址：")
        sys.argv.append(f)
    x=""
    if  (len(sys.argv)==2):
        filename=str(sys.argv[1]).strip('"').strip("'")
        if not filename.endswith('.km') and  not filename.endswith('.zxm') and  not filename.endswith('.xmind'):
            print("指定文件不正确，是检查路径是否存在空格，&等特殊字符，如果存在，请在头尾添加双引号！\n")
            sys.exit(1)
        x=input("请填写需求id：")
        sys.argv.append(x)
    p=""
    if  (len(sys.argv)==3):
        pv=""
        
        if os.environ.get('HOME') is not None:
            pv=os.path.join(os.environ.get('HOME'),".km")
        elif os.environ.get('HOMEPATH') is not None:
            pv=os.path.join(os.environ.get('HOMEPATH'),".km")
        if len(pv)>0:
            if os.path.exists(pv):
                for i in os.listdir(pv):
                    if i.startswith('pzhanghaofile.'):
                        p=i[14:]
                        break
        if len(p)==0:
            p=input("请填写p账号：")
            try:
                os.makedirs(pv)
                with open(os.path.join(pv,'pzhanghaofile.'+p),"w") as f1:
                    f1.write("\n")
            except:
                pass
        sys.argv.append(p)
    if len(sys.argv)==4:
        return filename ,sys.argv[2],sys.argv[3]
    else:
        print("输入的参数过多~",sys.argv)
        sys.exit(1)

# if __name__ == "__main__":
def main():
    run(*put())

    # print(*put())

def maine():
    # print(*put())
    run(*put(),False)

        

   
def mainInit():
    print(sys.argv,sys.argv[0][:-3])
#    os.system('''reg add "HKEY_CLASSES_ROOT\SystemFileAssociations\.km\shell\km2Excel\command" /f /d "\\\"{}kme.exe\\\" \\\"%1\\\""'''.format(sys.argv[0][:-3]))
#    os.system('''reg add "HKEY_CLASSES_ROOT\SystemFileAssociations\.zxm\shell\zxm2Excel\command" /f /d "\\\"{}kme.exe\\\" \\\"%1\\\""'''.format(sys.argv[0][:-3]))
#    os.system('''reg add "HKEY_CLASSES_ROOT\SystemFileAssociations\.xmind\shell\xmind2Excel\command" /f /d "\\\"{}kme.exe\\\" \\\"%1\\\""'''.format(sys.argv[0][:-3]))
    
    os.system(r'''reg add "HKEY_CLASSES_ROOT\SystemFileAssociations\.km\shell\km2Excel\command" /f /d "\"{}kme.exe\" \"%1\""'''.format(sys.argv[0][:-3]))
    os.system(r'''reg add "HKEY_CLASSES_ROOT\SystemFileAssociations\.zxm\shell\zxm2Excel\command" /f /d "\"{}kme.exe\" \"%1\""'''.format(sys.argv[0][:-3]))
    os.system(r'''reg add "HKEY_CLASSES_ROOT\SystemFileAssociations\.xmind\shell\xmind2Excel\command" /f /d "\"{}kme.exe\" \"%1\""'''.format(sys.argv[0][:-3]))


    # pv=""
    # p=""
    # if os.environ.get('HOME') is not None:
    #     pv=os.path.join(os.environ.get('HOME'),".km")
    # elif os.environ.get('HOMEPATH') is not None:
    #     pv=os.path.join(os.environ.get('HOMEPATH'),".km")
    # if len(pv)>0:
    #     if os.path.exists(pv):
    #         for i in os.listdir(pv):
    #             if i.startswith('pzhanghaofile.'):
    #                 p=i[14:]
    #                 break
    # if len(p)==0:
    #     p=input("请填写p账号：")
    #     try:
    #         os.makedirs(pv)
    #         with open(os.path.join(pv,'pzhanghaofile.'+p),"w") as f1:
    #             f1.write("\n")
    #     except:
    #         pass
# maine()
# def un_zip(file_name):
#     """unzip zip file"""
#     zip_file = zipfile.ZipFile(file_name)
#     data=zip_file.read('content.json').decode()
#     data=json.loads(data)
#     zip_file.close()
#     return data

# un_zip(r'C:\Users\Administrator\Envs\test1\test\腾讯视频增加畅读卡落地页.zxm')





        

    

