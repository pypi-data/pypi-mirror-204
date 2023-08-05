import os, copy
import tkinter as tk
from tkinter import (messagebox, filedialog, simpledialog)
from openpyxl import (load_workbook, Workbook)
import tksheet

colors=['#FFBBBB', '#BBFFBB', '#BBBBFF', '#CCCCFF']

class sheet:
    def saveCleanedData(self, evnt=None):
        wb=Workbook()
        ws = wb.active

        for r, row in enumerate(self.shtL.get_sheet_data()):
            for c, v in enumerate(row):
                ws[f'{chr(65+c)}{r+1}']=v

        fn = self.title.split('.xlsx')[0]+'_cleaned.xlsx'
        try:
            wb.save(fn)
            self.hL.config(text=f'{fn} 저장 완료.', fg='green')
        except PermissionError:
            self.hL.config(text=f'{fn}가 열려있어 저장불가', fg='red')


    def combineRowsColumns(self, evnt=None):
        if not self.shtL.anything_selected(): return
        rowFrom, colFrom, rowTo, colTo = self.shtL.get_all_selection_boxes()[0]
        data = self.shtL.get_sheet_data()
        self.dataBakup=copy.deepcopy(data)

        if rowTo-rowFrom < len(data) and colTo-colFrom == len(data[0]):
            self.actionDone=['combineRows', [rowFrom, rowTo]]
            row=[]
            for j in range(len(data[rowFrom])):
                tmp=[s for s in [d[j].strip() if isinstance(d[j], str) else str(d[j]) for d in data[rowFrom:rowTo]] if s != 'None']
                row.append(' '.join(tmp))

            data=data[:rowFrom]+[row]+data[rowTo:]
        elif rowTo-rowFrom == len(data) and colTo-colFrom < len(data[0]):
            self.actionDone=['combineColumns', [colFrom, colTo]]

            for row, dat in enumerate(data):
                tmp=[s for s in [s.strip() if isinstance(s, str) else str(s) for s in dat[colFrom:colTo]] if s != 'None']
                data[row]=dat[:colFrom]+[' '.join(tmp)]+dat[colTo:]

        self.shtL.set_sheet_data(data)
        self.checkSimilars()
        self.highlightRows()


    def addCells(self, evnt=None):
        if not self.shtL.anything_selected(): return
        rowFrom, col, _, _ = self.shtL.get_all_selection_boxes()[0] # 한 행만 선택
        data = self.shtL.get_sheet_data()
        self.dataBakup=copy.deepcopy(data)
        self.actionDone=['add', [col, rowFrom]]

        for row in data:
            row.append(None)
        data[rowFrom].pop()
        data[rowFrom].insert(col, None)

        data[0][-1]='Xtra'
        self.shtL.set_sheet_data(data)
        self.checkSimilars()
        self.highlightRows()



    def subCells(self, evnt=None):
        if not self.shtL.anything_selected(): return
        rowFrom, col, _, _ = self.shtL.get_all_selection_boxes()[0]

        data = self.shtL.get_sheet_data()
        self.dataBakup=copy.deepcopy(data)
        self.actionDone=['sub', [col, rowFrom]]
        data[rowFrom].append(data[rowFrom].pop(col))
        
        self.shtL.set_sheet_data(data)
        self.checkSimilars()
        self.highlightRows()
        


    def delRowsColumns(self, evnt=None):
        if not self.shtL.anything_selected(): return
        rowFrom, colFrom, rowTo, colTo = self.shtL.get_all_selection_boxes()[0]

        data = self.shtL.get_sheet_data()
        self.dataBakup=copy.deepcopy(data)

        if rowTo-rowFrom < len(data) and colTo-colFrom == len(data[0]):
            self.actionDone=['delRows', [rowFrom, rowTo]]
            data=data[:rowFrom]+data[rowTo:]
        elif rowTo-rowFrom == len(data) and colTo-colFrom < len(data[0]):
            self.actionDone=['delColumns', [colFrom, colTo]]
            data=[row[:colFrom]+row[colTo:] for row in data]

        self.shtL.set_sheet_data(data)
        self.checkSimilars()
        self.highlightRows()

        
        
    def repeat(self, evnt=None):
        data = self.shtL.get_sheet_data()
        if self.actionDone[0] == 'delRows':
            rowsN, rows2delete=self.actionDone[2]
            if rowsN > 1:
                for j in rows2delete[::-1]:
                    data=data[:j]+data[j+rowsN:]
            else:
                for j in rows2delete[::-1]:
                    data.pop(j)
            
        elif self.actionDone[0] == 'sub':
            col, rows2delete=self.actionDone[2]
            for j in rows2delete:
                data[j].append(data[j].pop(col))

        elif self.actionDone[0] == 'add':
            col, rows2add=self.actionDone[2]
            for j in rows2add:
                data[j].insert(col, None)
                data[j].pop()

        self.shtL.set_sheet_data(data)
            

    
    def checkSimilars(self):
        data = self.shtL.get_sheet_data()
        flag2Repeat=False
        if self.actionDone[0] == 'delRows':
            rowFrom, rowTo = self.actionDone[1]
            if rowTo-rowFrom > 1: # 여러 행이 삭제된 경우,  삭제된 행의 첫번째 열들로 구성된 리스트..
                deletedRows=[row[0] for row in self.dataBakup[rowFrom:rowTo]]
                rowsN=len(deletedRows)
                rows2delete=[]
                for j in range(len(data)):
                    if  [row[0] for row in data[j:j+rowsN]] == deletedRows:
                        rows2delete.append(j)
            else: # 오직 1 행만 삭제한 경우, 행과 동일한 경우...
                deletedRows=self.dataBakup[rowFrom]
                rowsN=1
                rows2delete=[]
                for j, row in enumerate(data):
                    if  row == deletedRows:
                        rows2delete.append(j)
                
            if rows2delete:
                self.actionDone.append((rowsN, rows2delete))
                flag2Repeat=True

        elif self.actionDone[0] == 'sub':
            col, rowFrom = self.actionDone[1]

            deletedRowTypes=[self.getType(item) for item in self.dataBakup[rowFrom]]
            rows2delete=[]
            for j, row in enumerate(data[1:]):
                if [self.getType(item) for item in row] == deletedRowTypes:
                    rows2delete.append(j+1)
                    
            if rows2delete:
                self.actionDone.append((col, rows2delete))
                flag2Repeat=True
            
        elif self.actionDone[0] == 'add':
            col, rowFrom = self.actionDone[1]

            addedRowTypes=[self.getType(item) for item in self.dataBakup[rowFrom]]+[None]
            rows2add=[]
            for j, row in enumerate(data[1:]):
                if [self.getType(item) for item in row] == addedRowTypes:
                    rows2add.append(j+1)

            if rows2add:
                self.actionDone.append((col, rows2add))
                flag2Repeat=True

        if flag2Repeat: self.repeat()


    def highlightRows(self):
        self.shtL.set_all_cell_sizes_to_text(redraw = True)
        rows = self.shtL.get_sheet_data()

        kinds=[]
        for row in rows[1:]:
            tmp=[]
            for item in row:
                if item is None:
                    kind='none'
                elif isinstance(item, (int, float)):
                    kind='num'
                elif isinstance(item, str):
                    if item.isalpha():
                        kind='alpha'
                    elif item.replace('.','').isnumeric():
                        kind='num'
                    else:
                        kind = 'alnum'
                else:
                    kind='???'
                tmp.append(kind)
            kinds.append(tuple(tmp))
            
        tmp={kind: kinds.count(kind) for kind in set(kinds)}
        patterns={kind: count for kind, count in sorted(tmp.items(), key=lambda t: t[1])}

        patternlist=iter(patterns.keys())
        indices0=[j for j, p in enumerate(next(patternlist)) if p == 'none']
        
        while indices0 and patternlist:
            try:
                patt = next(patternlist)
            except:
                break

            indicesJ =[j for j, p in enumerate(patt) if p == 'none']
            for v in copy.deepcopy(indices0):
                if v not in indicesJ: indices0.remove(v)

        self.shtL.dehighlight_all()
        if indices0:
            #print('deleting column of  ', indices0[-1])
            self.shtL.select_column(column=indices0[-1], redraw=True)
            self.delRowsColumns()
        else:
            indices = [j+1 for j, kind in enumerate(kinds) if kind == list(patterns.keys())[0]]
            self.shtL.highlight_rows(rows = indices, bg = colors[2], redraw = True)
            self.shtL.select_row(row=indices[0], redraw=True)
            self.shtL.see(row=indices[0])
                
            
        self.saveCleanedData()
        rows = self.shtL.get_sheet_data()
        if self.hL['fg'] == 'green':
            self.hL.config(text=f'총 항목 수: {len(rows[0])}×{len(rows)}   총 줄-패턴 종류 수: {len(patterns)}')



    @staticmethod
    def getType(a):
        if isinstance(a, str):
            if a.isnumeric() or ('.' in a and a.replace('.','').isnumeric()): return 'num'
            else: return 'str'
        elif isinstance(a, int):
            return 'int'
        elif isinstance(a, float):
            return 'float'
        

    def restore(self, event=None):
        if not self.dataBakup: return
        self.shtL.set_sheet_data(self.dataBakup)
        self.shtL.set_all_cell_sizes_to_text(redraw = True)
        self.data=self.dataBakup
        self.dataBakup=None
        self.actionDone=None
        self.hL.config(text=f'총 항목 수: {len(self.data[0])}×{len(self.data)}')



    def __init__(self, xlpath: str):

        self.dataBakup=None
        self.actionDone=None
        
        self.title=xlpath
        if os.path.exists(xlpath):
            wb=load_workbook(xlpath, data_only=True)
            data = [[f.value for f in row] for row in wb.worksheets[0].rows]
            self.data=[row for row in data if row != [None for _ in range(len(data[0]))]]
        
        self.hw=tk.Tk()
        self.hw.option_add('*Font', ('Palatino Linotype',12))
        self.hw.geometry('1000x850')
        self.hw.title(self.title+' '*10+'F1: insert cell, F3: delete cell, F5: combine cols/rows, F7: delete cols/rows, F9: save, F12: restore')
        self.hw.bind('<F1>', self.addCells)
        self.hw.bind('<F3>', self.subCells)
        self.hw.bind('<F5>', self.combineRowsColumns)
        self.hw.bind('<F7>', self.delRowsColumns)
        self.hw.bind('<F9>', self.saveCleanedData)
        self.hw.bind('<F12>', self.restore)
        
        self.hL=tk.Label(self.hw, text='...')
        self.hL.pack(pady=2)

        self.shtL = tksheet.Sheet(self.hw)
        self.shtL.pack(expand=1, fill='both')
        self.shtL.enable_bindings()
        self.shtL.disable_bindings(['rc_delete_row', 'rc_delete_column'])
        self.shtL.disable_bindings(['rc_delete_row', 'rc_delete_column', 'cut', 'paste', 'copy', 'delete'])        

        self.shtL.popup_menu_add_command("insert cell (F1)", self.addCells)
        self.shtL.popup_menu_add_command("delete cell (F3)", self.subCells)
        self.shtL.popup_menu_add_command("combine rows/cols (F5)", self.combineRowsColumns)
        self.shtL.popup_menu_add_command("delete rows/cols (F7)", self.delRowsColumns)
        self.shtL.popup_menu_add_command("restore (F12)", self.restore)

        self.shtL.set_sheet_data(self.data)
        self.shtL.set_all_cell_sizes_to_text(redraw = True)

        self.hL.config(text=f'총 항목 수: {len(self.data[0])}×{len(self.data)}')
        self.hw.mainloop()        
    


if __name__ == '__main__':
    path=r'../AJ-SO007558.xlsx'
    sheet(path)
